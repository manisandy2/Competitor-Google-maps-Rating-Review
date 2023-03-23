from selenium import webdriver
from openpyxl import Workbook,load_workbook
from selenium.webdriver.common.by import By
import datetime

date = datetime.datetime.now().strftime("%d-%m-%Y")
driver = webdriver.Chrome(executable_path=r"D:\projects\competitor_maps_review\driver\chromedriver.exe")

wb = load_workbook(r"D:\projects\competitor_maps_review\excel\Competitor\Competitor\myg mobiles.xlsx")

ws = wb.active

for r in range(2,ws.max_row+1):
    try:
        print(ws.cell(row=r,column=2).value)
        driver.get(ws.cell(row=r,column=2).value)
        print(driver.find_element(By.CLASS_NAME, "aMPvhf-fI6EEc-KVuj8d").text)
        ws.cell(row=r,column=3).value = driver.find_element(By.CLASS_NAME, "aMPvhf-fI6EEc-KVuj8d").text
        print(driver.find_element(By.CLASS_NAME, "Yr7JMd-pane-hSRGPd").text[0:-8])
        ws.cell(row=r,column=4).value = driver.find_element(By.CLASS_NAME, "Yr7JMd-pane-hSRGPd").text[0:-8]

        wb.save(r"D:\projects\competitor_maps_review\save data\myg Mobiles report " + date +".xlsx")
    except:
        pass