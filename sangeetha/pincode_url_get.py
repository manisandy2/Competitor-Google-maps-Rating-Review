from selenium import webdriver
import time
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
# import datetime

# date = datetime.date.today().strftime("%#d-%#m-%Y")

driver = webdriver.Chrome(executable_path=r"D:\projects\competitor_maps_review\driver\chromedriver.exe")

wb = load_workbook(r"D:\projects\competitor_maps_review\sangeetha\sangeetha-mobiles-store-list-new.xlsx")

ws = wb.active
url = "https://www.google.co.in/maps?hl=en"
for r in range(2,ws.max_row+1):
    # print(ws.cell(row=r,column=4).value)
    pincode_code = "Sangeetha Mobiles Pvt Ltd " + str(ws.cell(row=r,column=4).value)
    print(pincode_code)
    driver.get(url)
    driver.implicitly_wait(5)
    driver.find_element(By.CLASS_NAME,"tactile-searchbox-input").send_keys(pincode_code)
    time.sleep(5)
    for name in driver.find_elements(By.CLASS_NAME,"DAdBuc"):
        print(name.text)
        name.find_element_by_class_name("ZHeE1b-c8csvc").click()
        time.sleep(5)






# for r in range(2, ws.max_row+1):
#     print(r)
#     pincode = ws.cell(row=r, column=2).value
#     print(pincode)
#     if search_pincode != pincode:
#         driver.get(url)
#         driver.implicitly_wait(5)
#         name = driver.find_element(By.CLASS_NAME,"tactile-searchbox-input").send_keys(search_pincode)
#         directions = driver.find_element(By.ID,"xoLGzf-T3iPGc").click()
#         driver.implicitly_wait(5)
#
#         nexts = driver.find_element_by_id("sb_ifc51").find_element_by_tag_name("input").send_keys(pincode)
#         time.sleep(1)
#         button_key = driver.find_element_by_id("sb_ifc51").find_element_by_tag_name("input").send_keys(Keys.ENTER)
#
#         time.sleep(1)
#         l = 3
#         for selection in driver.find_elements(By.CLASS_NAME,"xB1mrd-T3iPGc-iSfDt-ij8cu"):
#             # print(selection.text)
#             for km in selection.find_elements(By.CLASS_NAME,"xB1mrd-T3iPGc-iSfDt-tUvA6e"):
#                 print(km.text)
#
#                 ws.cell(row=r,column=l).value = km.text
#                 l = l + 1
#                 wb.save(r"D:\mani\distance5_10_km\Save Data\TN-T.NAGAR 2-" + date + ".xlsx")
        # ws.cell(row=r,column=l).value = n.find_element_by_class_name("section-directions-trip-secondary-text").text

        # wb.save(r"D:\mani\distance5_10_km\Save Data\TN-T.NAGAR 2-" + date + ".xlsx")
#     elif search_pincode == pincode:
#         ws.cell(row=r, column=4).value = "0 km"
#         wb.save(r"D:\mani\distance5_10_km\Save Data\Tn-Ekkattuthangal.xlsx")
# driver.quit()
#
# wb1 = load_workbook(r"D:\mani\distance5_10_km\Save Data\Tn-Ekkattuthangal.xlsx")
#
# for ws1 in wb1:
#     # print(ws.title)
#     if ws1.title == "pincode":
#         break
#
# wb1.active
#
# for r1 in range(2, ws1.max_row+1):
#     list_km = ws1.cell(row=r1, column=4).value
#     print(list_km[0:-5])
#     # print(list_km)
#     if float(list_km[0:-3]) <= float(5):
#         ws1.cell(row=r1, column=5).value = list_km
#     elif float(list_km[0:-3]) <= 10:
#         ws1.cell(row=r1, column=6).value = list_km
#     elif float(list_km[0:-3]) <= 15:
#         ws1.cell(row=r1, column=7).value = list_km
#     else:
#         ws1.cell(row=r1, column=8).value = list_km
#
#
# wb1.save(r"D:\mani\distance5_10_km\Save Data\Tn-Ekkattuthangal final.xlsx")