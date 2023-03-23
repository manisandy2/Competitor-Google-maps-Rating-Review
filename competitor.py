from selenium import webdriver
from openpyxl import Workbook,load_workbook
from selenium.webdriver.common.by import By
import datetime
import pandas as pd
date = datetime.datetime.now().strftime("%d-%m-%Y")
driver = webdriver.Chrome(executable_path=r"D:\projects\competitor_maps_review\driver\chromedriver.exe")


class GoogleMaps:
    def __init__(self):
        pass

    def google_report(self, **kwargs):
        print(kwargs.get('link'))
        print(kwargs.get('col'))
        print(kwargs.get('col')+1)
        print(kwargs.get('title'))

        wb = load_workbook(kwargs.get('link'))
        ws = wb.active
        for r in range(2,ws.max_row+1):
            try:
                print(r)
                print(ws.cell(row=r,column=kwargs.get('col')).value)
                driver.get(ws.cell(row=r,column=kwargs.get('col')).value)
                print(driver.find_element(By.CLASS_NAME, "aMPvhf-fI6EEc-KVuj8d").text)
                ws.cell(row=r,column=kwargs.get('col')+1).value = driver.find_element(By.CLASS_NAME, "aMPvhf-fI6EEc-KVuj8d").text
                print(driver.find_element(By.CLASS_NAME, "Yr7JMd-pane-hSRGPd").text[0:-8])
                ws.cell(row=r,column=kwargs.get('col')+2).value = driver.find_element(By.CLASS_NAME, "Yr7JMd-pane-hSRGPd").text[0:-8]
                s = r"D:\projects\competitor_maps_review\save data\ " + kwargs.get('title') + date + " report.xlsx"

                wb.save(s)
            except:
                pass